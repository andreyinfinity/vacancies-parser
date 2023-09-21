"""
Модуль работы с файлами. Содержит классы для работы с файлами Json.
"""
from abc_classes import AbsFile
import json
from openpyxl import Workbook
import os


class JsonFile(AbsFile):
    def __init__(self, filename: str):
        self.filename = filename

    def save_to_file(self, json_data: json) -> None:
        """
        Метод сохранения в JSON файл.
        :param json_data:
        :return:
        """
        os.makedirs(os.path.split(self.filename)[0], exist_ok=True)
        with open(self.filename, "w", encoding="utf-8") as json_file:
            json.dump(json_data, json_file, indent=2, ensure_ascii=False)

    def load_from_file(self) -> json:
        with open(self.filename, "r", encoding="utf-8") as json_file:
            json_data = json.load(json_file)
        return json_data

    def find_in_list(self, _list: list, val: str, key_name: str, result=None) -> list:
        """
        Поиск по значению ключа в списке словарей. Находит город и его id.
        :param val: Слово для поиска(город)
        :param key_name: Название ключа в словаре, значение которого задается в поиске
        :param _list:
        :param result: Список для записи найденных значений
        :return:
        """

        if result is None:
            result = []
        for item in _list:
            self.find_in_dict(_dict=item, val=val, key_name=key_name, result=result)
        return result

    def find_in_dict(self, _dict: dict, val: str, key_name: str, result: list):
        if not _dict.get('areas'):
            if _dict.get(key_name).lower() == val.lower():
                result.extend([int(_dict.get('id')), _dict.get(key_name)])
                return result
        else:
            self.find_in_list(_list=_dict.get('areas'), val=val, key_name=key_name, result=result)


class ExelFile(AbsFile):
    """Класс для работы с файлами Exel. При инициализации передается путь к файлу."""
    def __init__(self, filename: str):
        self.filename = filename

    def save_to_file(self, json_data: list[dict]) -> None:
        """Преобразует json формат в таблицу Exel и сохраняет в файл."""
        os.makedirs(os.path.split(self.filename)[0], exist_ok=True)
        workbook = Workbook()
        # создание листа книги
        sheet = workbook.create_sheet('Вакансии', 0)
        # создание заголовков
        col_num = 1
        for key in json_data[0].keys():
            sheet.cell(row=1, column=col_num, value=key)
            col_num += 1
        # Запись значений в ячейки
        row_num = 2
        for item in json_data:
            col_num = 1
            for value in item.values():
                sheet.cell(row=row_num, column=col_num, value=value)
                col_num += 1
            row_num += 1
        # Сохранение файла
        workbook.save(self.filename)

    def load_from_file(self):
        pass


if __name__ == "__main__":
    from config import XLSX_FILE, VAC_FILE
    js = JsonFile(VAC_FILE)
    data = js.load_from_file()

    file = ExelFile(XLSX_FILE)
    file.save_to_file(data)
    # hh_areas = HH_AREAS
    # hh_file = JsonFile(hh_areas)
    # area = hh_file.load_from_file()
    # print(JsonFile.find_in_list(area, "казань", 'name'))

    # sj_areas = SJ_AREAS
    # sj_file = JsonFile(sj_areas)
    # area = sj_file.load_from_file()
    # print(JsonFile.find_in_list(area["objects"], "казань", 'title'))
